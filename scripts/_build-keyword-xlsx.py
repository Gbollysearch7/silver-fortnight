import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

kws = json.load(open('data/planned-keywords.json'))

wb = Workbook()

# ---- shared styles ----
hdr_fill = PatternFill('solid', fgColor='1F2A44')
hdr_font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
title_font = Font(bold=True, size=16, name='Calibri', color='1F2A44')
sub_font = Font(italic=True, size=10, color='666666')
bold = Font(bold=True, name='Calibri')
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin,right=thin,top=thin,bottom=thin)
wave_fills = {1:'E8F2EC', 2:'FBF3E2', 3:'F6E8E2'}

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(vertical='center', horizontal='left')
        cell.border = border

# ============ TAB 1: SUMMARY / CADENCE / KPIs ============
ws = wb.active
ws.title = 'Plan & KPIs'
ws.sheet_view.showGridLines = False
ws['A1'] = 'TradersYard SEO — Content Execution Plan'; ws['A1'].font = title_font
ws['A2'] = 'Planned keyword pipeline, posting cadence, and KPI targets  |  Source: SEMrush + Google Search Console'; ws['A2'].font = sub_font

r = 4
ws.cell(r,1,'THE PLAN AT A GLANCE').font = Font(bold=True, size=12, color='1F2A44')
r += 1
glance = [
    ('Total planned pages (net-new)', '676'),
    ('Total monthly search volume targeted', '58,330 / month'),
    ('Posting cadence', '3 posts per week'),
    ('Posts per month', '~12'),
    ('Posts per year', '~144'),
    ('Approach', 'SEO Avalanche — easiest keywords first, building authority before harder terms'),
]
for k,v in glance:
    ws.cell(r,1,k).font = bold; ws.cell(r,2,v)
    r += 1

r += 1
ws.cell(r,1,'BUILD WAVES (easiest first)').font = Font(bold=True, size=12, color='1F2A44')
r += 1
wave_hdr = ['Wave', 'Difficulty', 'Pages', 'Monthly volume', 'Est. timeline at 3/week']
for i,h in enumerate(wave_hdr, 1): ws.cell(r,i,h)
style_header(ws, r, len(wave_hdr)); r += 1
wave_rows = [
    ['Wave 1 — Quick Wins', 'KD 0-10 (very low)', 318, '18,520', 'Months 1-26'],
    ['Wave 2 — Easy', 'KD 11-20 (low)', 210, '18,560', 'Months 27-44'],
    ['Wave 3 — Medium', 'KD 21-30 (moderate)', 148, '21,250', 'Months 45-56'],
]
for row in wave_rows:
    for i,v in enumerate(row,1):
        cell = ws.cell(r,i,v); cell.border = border
    r += 1
ws.cell(r,1,'Note: cadence is sustainable and can be increased as results compound. Wave 1 front-loads the fastest, highest-probability wins.').font = sub_font
r += 2

ws.cell(r,1,'KPI TARGETS').font = Font(bold=True, size=12, color='1F2A44')
r += 1
ws.cell(r,1,'Baseline is the current 28-day Google Search Console figures. Targets are cumulative as new content ranks.').font = sub_font
r += 1
kpi_hdr = ['KPI', 'Baseline (now)', '3-month target', '6-month target', '12-month target']
for i,h in enumerate(kpi_hdr,1): ws.cell(r,i,h)
style_header(ws, r, len(kpi_hdr)); r += 1
kpi_rows = [
    ['Search impressions / month', '55,900', '70,000', '95,000', '160,000'],
    ['Blog clicks / month', '1,560', '2,200', '3,500', '7,000'],
    ['Distinct queries ranking', '1,209', '1,500', '2,000', '3,200'],
    ['Blog pages ranking', '210', '250', '320', '450'],
    ['Average position', '12.1', '11.0', '9.5', '8.0'],
    ['Queries in positions 1-3', '86', '130', '220', '450'],
    ['New posts published (cumulative)', '0', '36', '72', '144'],
]
for row in kpi_rows:
    for i,v in enumerate(row,1):
        cell = ws.cell(r,i,v); cell.border = border
        if i==1: cell.font = bold
    r += 1
r += 1
ws.cell(r,1,'Targets are directional and assume the 3-posts-per-week cadence plus the title/header rewrite programme on existing pages. SEO results compound, so growth accelerates in later months.').font = sub_font

# widths
for col,w in zip('ABCDE',[34,18,16,16,18]): ws.column_dimensions[col].width = w

# ============ TAB 2: WEEKLY/MONTHLY SCHEDULE ============
ws2 = wb.create_sheet('Posting Schedule')
ws2.sheet_view.showGridLines = False
ws2['A1'] = 'Posting Schedule — 3 posts per week'; ws2['A1'].font = title_font
ws2['A2'] = 'A predictable weekly rhythm. Each post is keyword-targeted, fact-checked, and quality-reviewed before publishing.'; ws2['A2'].font = sub_font
r = 4
ws2.cell(r,1,'WEEKLY RHYTHM').font = Font(bold=True, size=12, color='1F2A44'); r+=1
sched_hdr = ['Day', 'Activity', 'Output']
for i,h in enumerate(sched_hdr,1): ws2.cell(r,i,h)
style_header(ws2,r,3); r+=1
week = [
    ['Monday', 'Publish post #1 (highest-volume keyword of the week)', '1 live article'],
    ['Tuesday', 'Keyword research + draft next posts', 'Drafts in pipeline'],
    ['Wednesday', 'Publish post #2', '1 live article'],
    ['Thursday', 'Internal linking + title/meta rewrites on existing pages', 'On-page improvements'],
    ['Friday', 'Publish post #3 + weekly performance check (GSC)', '1 live article + report'],
]
for row in week:
    for i,v in enumerate(row,1):
        c=ws2.cell(r,i,v); c.border=border
        if i==1: c.font=bold
    r+=1
r+=1
ws2.cell(r,1,'MONTHLY RHYTHM (~12 posts)').font = Font(bold=True, size=12, color='1F2A44'); r+=1
mhdr=['Week','Focus','Posts']
for i,h in enumerate(mhdr,1): ws2.cell(r,i,h)
style_header(ws2,r,3); r+=1
month=[
    ['Week 1','Wave 1 quick-win keywords','3'],
    ['Week 2','Wave 1 quick-win keywords','3'],
    ['Week 3','Wave 1 + 1 cluster/pillar piece','3'],
    ['Week 4','Wave 1 + monthly performance review','3'],
]
for row in month:
    for i,v in enumerate(row,1):
        c=ws2.cell(r,i,v); c.border=border
        if i==1: c.font=bold
    r+=1
for col,w in zip('ABC',[14,58,22]): ws2.column_dimensions[col].width=w

# ============ TAB 3: FULL KEYWORD LIST ============
ws3 = wb.create_sheet('Planned Keywords')
ws3.sheet_view.showGridLines = False
ws3['A1'] = 'Planned Keywords — full pipeline (676 pages)'; ws3['A1'].font = title_font
ws3['A2'] = 'Sorted by wave (easiest first), then by monthly search volume. KD = keyword difficulty (lower is easier).'; ws3['A2'].font = sub_font
r=4
khdr=['#','Keyword','Monthly volume','KD','Intent','Content pillar','Wave']
for i,h in enumerate(khdr,1): ws3.cell(r,i,h)
style_header(ws3,r,len(khdr))
ws3.freeze_panes = ws3.cell(r+1,1)
r+=1
for idx,k in enumerate(kws,1):
    vals=[idx,k['kw'],k['vol'],k['kd'],k.get('intent',''),k['pillar'],'Wave '+str(k['wave'])]
    fill = wave_fills.get(k['wave'])
    for i,v in enumerate(vals,1):
        c=ws3.cell(r,i,v); c.border=border
        if fill and i==7: c.fill = PatternFill('solid', fgColor=fill)
    r+=1
for col,w in zip('ABCDEFG',[6,46,16,8,26,28,10]): ws3.column_dimensions[col].width=w
ws3.auto_filter.ref = f"A4:G{r-1}"

wb.save('TradersYard-SEO-Keyword-Plan.xlsx')
print('Saved TradersYard-SEO-Keyword-Plan.xlsx with', len(kws), 'keywords across 3 tabs')
